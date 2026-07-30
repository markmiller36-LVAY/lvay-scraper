#!/usr/bin/env python3
"""Build the official, season-scoped 2026-27 LHSAA alignment dataset."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook


SPORT_KEYS = {
    "Baseball": "baseball",
    "Boys Basketball": "boys_basketball",
    "Girls Basketball": "girls_basketball",
    "Football": "football",
    "Boys Soccer": "boys_soccer",
    "Girls Soccer": "girls_soccer",
    "Softball": "softball",
    "Volleyball": "volleyball",
}

SEASON_KEYS = {
    "football": 2026,
    "volleyball": 2026,
    "boys_basketball": 2027,
    "girls_basketball": 2027,
    "boys_soccer": 2027,
    "girls_soccer": 2027,
    "baseball": 2027,
    "softball": 2027,
}

CLASS_ROW = re.compile(r"^(5A|4A|3A|2A|1A|B|C)\s+(\d+)\s+(.+)$")
DIVISION_DISTRICT_ROW = re.compile(
    r"^(I|II|III|IV|V)\s+(\d+)\s+(.+)$"
)
DIVISION_ROW = re.compile(r"^(I|II|III|IV|V)\s+(.+)$")


def clean_school_name(value: str) -> str:
    """Remove LHSAA report legend markers without changing the real name."""
    return re.sub(r"\s+[+*]+$", "", value).strip()


def alignment_status(value: str) -> str:
    if re.search(r"\s+\*+$", value):
        return "jv_only"
    if re.search(r"\s+\++$", value):
        return "non_district_honors"
    return "standard"


def normalized_name(value: str) -> str:
    value = clean_school_name(value)
    value = (
        value.replace("\u2019", "'")
        .replace("\u2018", "'")
        .replace("\u2013", "-")
        .replace("\u2014", "-")
        .replace("&", "and")
    )
    return re.sub(r"\s+", " ", value).strip().casefold()


def load_workbook_assignments(path: Path) -> tuple[
    dict[str, str],
    dict[str, str],
    dict[str, dict[str, str]],
]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    exact: dict[str, str] = {}
    normalized: dict[str, str] = {}
    divisions = {
        "volleyball": {},
        "boys_soccer": {},
        "girls_soccer": {},
    }
    for row in sheet.iter_rows(
        min_row=2, values_only=True
    ):
        school, _enrollment, basic_class = row[:3]
        if not school or not basic_class:
            continue
        school_name = clean_school_name(str(school))
        class_name = str(basic_class).strip()
        exact[school_name] = class_name
        normalized_school = normalized_name(school_name)
        normalized[normalized_school] = class_name
        for sport, index in (
            ("volleyball", 7),
            ("boys_soccer", 8),
            ("girls_soccer", 9),
        ):
            value = row[index]
            if value and str(value).strip() != "-":
                divisions[sport][normalized_school] = (
                    f"Division {str(value).strip()}"
                )
    workbook.close()
    return exact, normalized, divisions


def parse_alignment_pdf(path: Path) -> tuple[dict, dict]:
    regular = {sport: {} for sport in SPORT_KEYS.values()}
    postseason = {
        sport: {}
        for sport in (
            "baseball",
            "boys_basketball",
            "girls_basketball",
            "football",
            "softball",
        )
    }

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            lines = [
                line.strip()
                for line in (page.extract_text() or "").splitlines()
                if line.strip()
            ]
            if len(lines) < 4 or not lines[0].startswith("2026 - 2027"):
                continue
            if not lines[1].startswith("Sport "):
                continue
            sport = SPORT_KEYS.get(lines[1][6:].strip())
            if not sport:
                continue
            is_postseason = "Post Season" in lines[0]

            for line in lines[3:]:
                if line.startswith(("* denotes", "+ denotes")):
                    continue
                if is_postseason:
                    match = DIVISION_ROW.match(line)
                    if match and sport in postseason:
                        division, school = match.groups()
                        postseason[sport][clean_school_name(school)] = (
                            f"Division {division}"
                        )
                    continue

                if sport in ("boys_soccer", "girls_soccer", "volleyball"):
                    match = DIVISION_DISTRICT_ROW.match(line)
                    if match:
                        division, district, school = match.groups()
                        regular[sport][clean_school_name(school)] = {
                            "division": f"Division {division}",
                            "district": int(district),
                            "alignment_status": alignment_status(school),
                        }
                else:
                    match = CLASS_ROW.match(line)
                    if match:
                        school_class, district, school = match.groups()
                        regular[sport][clean_school_name(school)] = {
                            "class": school_class,
                            "district": int(district),
                            "alignment_status": alignment_status(school),
                        }

    return regular, postseason


def build_dataset(pdf_path: Path, workbook_path: Path) -> dict:
    regular, postseason = parse_alignment_pdf(pdf_path)
    basic_exact, basic_normalized, workbook_divisions = (
        load_workbook_assignments(workbook_path)
    )
    unresolved_classes: dict[str, list[str]] = {}
    missing_postseason: dict[str, list[str]] = {}
    division_mismatches: dict[str, list[dict[str, str]]] = {}

    sports: dict[str, dict] = {}
    for sport, rows in regular.items():
        compiled: dict[str, dict] = {}
        for school, values in rows.items():
            record = dict(values)
            school_class = record.get("class")
            if not school_class:
                school_class = basic_exact.get(school)
                if not school_class:
                    school_class = basic_normalized.get(
                        normalized_name(school)
                    )
            status = record.get("alignment_status", "standard")
            if not school_class and status == "standard":
                unresolved_classes.setdefault(sport, []).append(school)
            record["class"] = school_class or "Unknown"

            if sport in workbook_divisions and status == "standard":
                workbook_division = workbook_divisions[sport].get(
                    normalized_name(school)
                )
                if workbook_division != record.get("division"):
                    division_mismatches.setdefault(sport, []).append(
                        {
                            "school": school,
                            "pdf": record.get("division", ""),
                            "workbook": workbook_division or "missing",
                        }
                    )

            if sport in postseason:
                division = postseason[sport].get(school)
                if not division:
                    normalized_postseason = {
                        normalized_name(name): value
                        for name, value in postseason[sport].items()
                    }
                    division = normalized_postseason.get(
                        normalized_name(school)
                    )
                if not division and school_class in ("B", "C"):
                    division = f"Class {school_class}"
                if not division and status != "standard":
                    division = "Not Postseason Eligible"
                if not division:
                    missing_postseason.setdefault(sport, []).append(school)
                else:
                    record["division"] = division

            division = record.get("division", "")
            record["track"] = (
                "small-school"
                if division.startswith("Class ")
                else "unassigned"
                if division == "Not Postseason Eligible"
                else "combined"
            )
            record["source"] = pdf_path.name
            compiled[school] = record
        sports[sport] = compiled

    if unresolved_classes or missing_postseason:
        problems = {
            "unresolved_classes": unresolved_classes,
            "missing_postseason": missing_postseason,
        }
        raise ValueError(json.dumps(problems, indent=2, sort_keys=True))

    return {
        "cycle": "2026-2028",
        "display_season": "2026-2027",
        "season_keys": SEASON_KEYS,
        "sources": [pdf_path.name, workbook_path.name],
        "source_precedence": (
            f"{pdf_path.name} controls sport-specific divisions and "
            "districts; the workbook supplies Basic Class."
        ),
        "workbook_division_differences": division_mismatches,
        "sports": sports,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--alignment-pdf", type=Path, required=True)
    parser.add_argument("--basic-class-workbook", type=Path, required=True)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("sport_alignments_2026_2027.json"),
    )
    args = parser.parse_args()

    dataset = build_dataset(
        args.alignment_pdf,
        args.basic_class_workbook,
    )
    args.output.write_text(
        json.dumps(dataset, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    counts = {
        sport: len(rows)
        for sport, rows in dataset["sports"].items()
    }
    print(f"Wrote {args.output}")
    print(json.dumps(counts, sort_keys=True))
    differences = dataset.get("workbook_division_differences", {})
    print(
        "Workbook division differences documented:",
        sum(len(rows) for rows in differences.values()),
    )


if __name__ == "__main__":
    main()
